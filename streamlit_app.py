import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import re
import io
import requests
from typing import List, Tuple, Dict
from PIL import Image
import tempfile
import os

def download_image(url: str) -> Tuple[bool, str, str]:
    """
    Download image from URL and return success status, file path, and error message.
    """
    try:
        # Clean up the URL
        url = url.strip()
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        temp_file.write(response.content)
        temp_file.close()
        
        # Verify it's a valid image
        try:
            with Image.open(temp_file.name) as img:
                # Convert to RGB if necessary and save as PNG
                if img.mode in ('RGBA', 'P'):
                    img = img.convert('RGB')
                img.save(temp_file.name, 'PNG')
            
            return True, temp_file.name, ""
        except Exception as img_error:
            os.unlink(temp_file.name)
            return False, "", f"Invalid image format: {str(img_error)}"
            
    except requests.exceptions.RequestException as e:
        return False, "", f"Download failed: {str(e)}"
    except Exception as e:
        return False, "", f"Unexpected error: {str(e)}"

def process_excel_with_images(file_content: bytes, insert_images: bool = True, max_image_size: int = 200) -> Tuple[bytes, List[Dict]]:
    """
    Process Excel file to either replace @IMAGE functions or insert actual images.
    """
    workbook = load_workbook(io.BytesIO(file_content))
    changes = []
    temp_files = []
    
    # Pattern to match =@IMAGE("link") or =@IMAGE('link')
    pattern = r'=@IMAGE\s*\(\s*["\']([^"\']+)["\']\s*\)'
    
    try:
        # Process each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Iterate through all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Check if cell contains @IMAGE function
                        match = re.search(pattern, cell.value, flags=re.IGNORECASE)
                        if match:
                            url = match.group(1)
                            original_value = cell.value
                            
                            if insert_images:
                                # Try to download and insert the image
                                success, temp_path, error_msg = download_image(url)
                                
                                if success:
                                    try:
                                        # Create Excel image object
                                        excel_img = ExcelImage(temp_path)
                                        
                                        # Resize image to fit in cell
                                        excel_img.width = min(excel_img.width, max_image_size)
                                        excel_img.height = min(excel_img.height, max_image_size)
                                        
                                        # Clear the cell content
                                        cell.value = ""
                                        
                                        # Add image to worksheet anchored to the cell
                                        excel_img.anchor = f"{cell.column_letter}{cell.row}"
                                        worksheet.add_image(excel_img)
                                        
                                        # Adjust row height and column width to accommodate image
                                        worksheet.row_dimensions[cell.row].height = max(
                                            worksheet.row_dimensions[cell.row].height or 15,
                                            excel_img.height * 0.75  # Excel uses points, images use pixels
                                        )
                                        worksheet.column_dimensions[cell.column_letter].width = max(
                                            worksheet.column_dimensions[cell.column_letter].width or 8,
                                            excel_img.width * 0.15  # Rough conversion
                                        )
                                        
                                        changes.append({
                                            'sheet': sheet_name,
                                            'cell': f"{cell.column_letter}{cell.row}",
                                            'original': original_value,
                                            'action': 'Image inserted',
                                            'url': url,
                                            'status': 'Success'
                                        })
                                        
                                        temp_files.append(temp_path)
                                        
                                    except Exception as e:
                                        # If image insertion fails, fall back to formula replacement
                                        cell.value = re.sub(pattern, r'=IMAGE("\1")', original_value, flags=re.IGNORECASE)
                                        changes.append({
                                            'sheet': sheet_name,
                                            'cell': f"{cell.column_letter}{cell.row}",
                                            'original': original_value,
                                            'action': 'Formula replaced (image insertion failed)',
                                            'url': url,
                                            'status': f'Error: {str(e)}'
                                        })
                                        if temp_path:
                                            temp_files.append(temp_path)
                                else:
                                    # If download fails, replace with regular IMAGE formula
                                    cell.value = re.sub(pattern, r'=IMAGE("\1")', original_value, flags=re.IGNORECASE)
                                    changes.append({
                                        'sheet': sheet_name,
                                        'cell': f"{cell.column_letter}{cell.row}",
                                        'original': original_value,
                                        'action': 'Formula replaced (download failed)',
                                        'url': url,
                                        'status': f'Error: {error_msg}'
                                    })
                            else:
                                # Just replace the formula
                                new_value = re.sub(pattern, r'=IMAGE("\1")', original_value, flags=re.IGNORECASE)
                                cell.value = new_value
                                changes.append({
                                    'sheet': sheet_name,
                                    'cell': f"{cell.column_letter}{cell.row}",
                                    'original': original_value,
                                    'action': 'Formula replaced',
                                    'url': url,
                                    'status': 'Success'
                                })
        
        # Save modified workbook to bytes
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        
        return output_buffer.getvalue(), changes
    
    finally:
        # Clean up temporary files
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except:
                pass

def main():
    st.set_page_config(
        page_title="Excel Image Processor",
        page_icon="🖼️",
        layout="wide"
    )
    
    st.title("🖼️ Excel Image Processor")
    st.markdown("""
    This app processes `=@IMAGE("link")` functions in Excel files. You can choose to:
    1. **Insert actual images** - Downloads and embeds images directly into Excel cells
    2. **Replace formulas only** - Converts `=@IMAGE()` to `=IMAGE()` functions
    """)
    
    # Settings
    st.sidebar.header("⚙️ Settings")
    
    insert_images = st.sidebar.radio(
        "Processing Mode:",
        ["Insert actual images into Excel", "Replace formulas only"],
        help="Choose whether to download and embed images or just fix the formula syntax"
    ) == "Insert actual images into Excel"
    
    if insert_images:
        max_image_size = st.sidebar.slider(
            "Maximum image size (pixels):",
            min_value=50,
            max_value=500,
            value=200,
            step=25,
            help="Images larger than this will be resized to fit"
        )
        
        st.sidebar.info("⚠️ **Note:** Downloading images may take some time depending on the number of images and their sizes.")
    else:
        max_image_size = 200
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xls'],
        help="Upload an Excel file containing @IMAGE functions"
    )
    
    if uploaded_file is not None:
        try:
            # Read the file content
            file_content = uploaded_file.read()
            
            # Process the file
            if insert_images:
                progress_text = "Downloading images and processing Excel file..."
                progress_bar = st.progress(0)
                progress_bar.progress(10)
            else:
                progress_text = "Processing Excel file..."
            
            with st.spinner(progress_text):
                if insert_images:
                    progress_bar.progress(50)
                modified_content, changes = process_excel_with_images(file_content, insert_images, max_image_size)
                if insert_images:
                    progress_bar.progress(100)
                    progress_bar.empty()
            
            # Display results
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.subheader("📈 Processing Results")
                if changes:
                    success_count = len([c for c in changes if c['status'] == 'Success'])
                    error_count = len(changes) - success_count
                    
                    if success_count > 0:
                        st.success(f"✅ Successfully processed {success_count} @IMAGE functions!")
                    if error_count > 0:
                        st.warning(f"⚠️ {error_count} items had errors (see details below)")
                    
                    # Show changes in an expandable section
                    with st.expander("View all changes", expanded=len(changes) <= 5):
                        for i, change in enumerate(changes, 1):
                            status_icon = "✅" if change['status'] == 'Success' else "❌"
                            st.write(f"**{status_icon} Change {i}:**")
                            st.write(f"- Sheet: `{change['sheet']}`")
                            st.write(f"- Cell: `{change['cell']}`")
                            st.write(f"- Action: {change['action']}")
                            st.write(f"- URL: `{change['url']}`")
                            st.write(f"- Status: {change['status']}")
                            st.write("---")
                else:
                    st.info("ℹ️ No @IMAGE functions found in the file.")
            
            with col2:
                st.subheader("📥 Download Processed File")
                if changes:
                    # Generate filename for download
                    original_name = uploaded_file.name
                    name_parts = original_name.rsplit('.', 1)
                    suffix = "_with_images" if insert_images else "_formulas_fixed"
                    if len(name_parts) == 2:
                        download_name = f"{name_parts[0]}{suffix}.{name_parts[1]}"
                    else:
                        download_name = f"{original_name}{suffix}"
                    
                    st.download_button(
                        label="📥 Download Processed Excel File",
                        data=modified_content,
                        file_name=download_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Click to download the processed Excel file"
                    )
                else:
                    st.info("No changes were made to download.")
            
            # Summary statistics
            if changes:
                st.subheader("📊 Summary")
                
                # Count changes per sheet and status
                sheet_stats = {}
                for change in changes:
                    sheet = change['sheet']
                    status = 'Success' if change['status'] == 'Success' else 'Error'
                    
                    if sheet not in sheet_stats:
                        sheet_stats[sheet] = {'Success': 0, 'Error': 0}
                    sheet_stats[sheet][status] += 1
                
                # Display summary
                summary_data = []
                for sheet, stats in sheet_stats.items():
                    summary_data.append({
                        "Sheet": sheet,
                        "Successful": stats['Success'],
                        "Errors": stats['Error'],
                        "Total": stats['Success'] + stats['Error']
                    })
                
                summary_df = pd.DataFrame(summary_data)
                st.dataframe(summary_df, use_container_width=True)
                
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            st.info("Please make sure you uploaded a valid Excel file (.xlsx or .xls)")
    
    # Help section
    with st.expander("ℹ️ Help & Information"):
        st.markdown("""
        **Processing Modes:**
        
        1. **Insert actual images into Excel:**
           - Downloads images from URLs in @IMAGE functions
           - Embeds the actual images directly into Excel cells
           - Automatically resizes images to fit
           - Adjusts row height and column width
           - Falls back to formula replacement if image download fails
        
        2. **Replace formulas only:**
           - Simply converts `=@IMAGE("url")` to `=IMAGE("url")`
           - No image downloading or embedding
           - Fast processing
        
        **Supported Formats:**
        - `.xlsx` (Excel 2007+)
        - `.xls` (Excel 97-2003)
        - Common image formats: PNG, JPG, JPEG, GIF, BMP
        
        **Image Processing Features:**
        - Automatic image resizing to specified maximum size
        - Row and column dimension adjustment
        - Error handling with fallback to formula replacement
        - Support for various image formats
        
        **Pattern Recognition:**
        - Matches: `=@IMAGE("link")`, `=@IMAGE('link')`, `= @IMAGE( "link" )`
        - Case insensitive matching
        - Handles various spacing variations
        
        **Notes:**
        - Image insertion requires internet connection to download images
        - Large files or many images may take longer to process
        - If an image fails to download, the cell will contain a regular IMAGE formula instead
        """)

if __name__ == "__main__":
    main()
